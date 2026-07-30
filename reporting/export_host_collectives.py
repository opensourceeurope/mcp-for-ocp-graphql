# /// script
# requires-python = ">=3.10"
# dependencies = ["fpdf2>=2.7", "openpyxl>=3.1"]
# ///
"""Export all collectives and funds hosted by a host (default: "europe" /
Open Source Europe), one row per account — built as input for categorizing
them. Funds matter: e.g. CaptainFact lives under this host as a FUND and
would be invisible to a COLLECTIVE-only query.

Per collective: creation date, last donation, last financial operation, a
Never Active flag (same definition as export_never_active_collectives.py —
no transactions incl. children, no published updates incl. children), the
long description (HTML stripped to plain text — clipped in md/pdf, full in
CSV), tags, social links, categories, and its children (events/projects
aggregated into one cell).

Multi-value cells (social links, children) hold one entry per line — children
as labeled name/description/tags blocks, a blank line between entries. For
spreadsheet use prefer --format xlsx: cells carry real line breaks with
wrap-text already applied, so they open multiline with zero post-processing
(CSV importers like Proton's drop in-cell newlines, so the CSV encodes them
as " ¶ " / " ¶¶ " markers instead; md/pdf tables flatten these cells to one
line).

No personal data is queried, so no token is needed — set OC_PERSONAL_TOKEN
only to raise rate limits. Admin contact data is a separate, deliberate step:
run export_collectives_admins.py with explicit slugs when you need it.

Usage, from inside the reporting/ directory (uv fetches the deps automatically
from the header above):

    uv run export_host_collectives.py                 # -> output/europe-collectives.md
    uv run export_host_collectives.py --format csv    # -> output/europe-collectives.csv
    uv run export_host_collectives.py --slug oce --format pdf
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from html.parser import HTMLParser

API_URL = "https://api.opencollective.com/graphql/v2"

# One page of hosted collectives (type COLLECTIVE). lastDonation is the newest
# incoming contribution, lastFinOp the newest financial operation of any kind.
# lastFinOpAll and the update lookups exist only to compute the "Never Active"
# flag with the same definition as export_never_active_collectives.py.
QUERY = """
query ($host: [AccountReferenceInput], $limit: Int!, $offset: Int!) {
  accounts(host: $host, type: [COLLECTIVE, FUND], limit: $limit, offset: $offset) {
    totalCount
    nodes {
      name
      slug
      isArchived
      createdAt
      longDescription
      tags
      socialLinks { type url }
      categories
      lastDonation: transactions(limit: 1, type: CREDIT, kind: [CONTRIBUTION]) {
        nodes { createdAt }
      }
      lastFinOp: transactions(limit: 1) {
        nodes { createdAt }
      }
      lastFinOpAll: transactions(limit: 1, includeChildrenTransactions: true) {
        nodes { createdAt }
      }
      lastUpdate: updates(limit: 1, onlyPublishedUpdates: true, orderBy: {field: PUBLISHED_AT, direction: DESC}) {
        nodes { publishedAt }
      }
      childrenAccounts(limit: 100) {
        nodes {
          name
          slug
          type
          description
          tags
          lastUpdate: updates(limit: 1, onlyPublishedUpdates: true, orderBy: {field: PUBLISHED_AT, direction: DESC}) {
            nodes { publishedAt }
          }
        }
      }
    }
  }
}
"""

# Each collective node carries nested transactions/updates/children lookups,
# so keep pages small or the API response gets slow enough to time out.
PAGE_SIZE = 25


def first_node(collection: dict) -> dict:
    nodes = (collection or {}).get("nodes") or []
    return nodes[0] if nodes else {}


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


def html_to_text(raw: str | None) -> str:
    """Strip HTML tags/entities and collapse whitespace to a single line."""
    if not raw:
        return ""
    parser = _TextExtractor()
    parser.feed(raw)
    return re.sub(r"\s+", " ", " ".join(parser.parts)).strip()


def format_social(links: list[dict] | None) -> str:
    """One "type: url" line per link, blank line between entries."""
    return "\n\n".join(
        f"{(link.get('type') or '').lower()}: {link.get('url')}"
        for link in links or []
        if link.get("url")
    )


def has_published_update(account: dict) -> bool:
    return bool(first_node(account.get("lastUpdate")))


def format_children(children: list[dict]) -> str:
    """One labeled block per child, blank line between children:

    name: Some Project (PROJECT)
    description: ...
    tags: a, b
    """
    blocks = []
    for child in children:
        lines = [f"name: {child.get('name') or child.get('slug') or ''} ({child.get('type') or ''})"]
        description = (child.get("description") or "").strip()
        if description:
            lines.append(f"description: {description}")
        tags = child.get("tags") or []
        if tags:
            lines.append(f"tags: {', '.join(tags)}")
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks)


def extract_row(collective: dict) -> dict:
    c_name = collective.get("name") or collective.get("slug") or ""
    c_slug = collective.get("slug") or ""
    if collective.get("isArchived"):
        c_name += " (archived)"

    donation = first_node(collective.get("lastDonation"))
    last_donation = (donation.get("createdAt") or "")[:10] or "never"
    fin_op = first_node(collective.get("lastFinOp"))
    last_fin_op = (fin_op.get("createdAt") or "")[:10] or "never"

    children = (collective.get("childrenAccounts") or {}).get("nodes") or []
    had_fin_op = bool(first_node(collective.get("lastFinOpAll")))
    had_update = has_published_update(collective) or any(
        has_published_update(child) for child in children
    )

    return {
        "collective": c_name,
        "collective_url": f"https://opencollective.com/{c_slug}",
        "created": (collective.get("createdAt") or "")[:10],
        "last_donation": last_donation,
        "last_fin_op": last_fin_op,
        "never_active": "no" if had_fin_op or had_update else "yes",
        "long_description": html_to_text(collective.get("longDescription")),
        "tags": ", ".join(collective.get("tags") or []),
        "social_links": format_social(collective.get("socialLinks")),
        "categories": ", ".join(collective.get("categories") or []),
        "children": format_children(children),
    }


COLUMNS = [
    ("collective", "Collective"),
    ("collective_url", "Collective URL"),
    ("created", "Created"),
    ("last_donation", "Last donation"),
    ("last_fin_op", "Last financial op"),
    ("never_active", "Never Active"),
    ("long_description", "Long Description"),
    ("tags", "Tags"),
    ("social_links", "Social Links"),
    ("categories", "Categories"),
    ("children", "Projects"),
]

# In Markdown output these text columns become clickable links to their URL
# column, and the raw-URL columns are dropped.
MD_LINKS = {"collective": "collective_url"}

# Full text only fits in CSV; md/pdf tables get these cells clipped (pdf much
# harder — its equal-width cells hold only a few dozen characters).
TRUNCATE_COLS = {"long_description", "children"}
MD_TRUNCATE_AT = 300
PDF_TRUNCATE_AT = 60


def clip(text: str, key: str, at: int) -> str:
    if key in TRUNCATE_COLS and len(text) > at:
        return text[: at - 1] + "…"
    return text


def flatten(text: str) -> str:
    """Collapse the multiline spreadsheet cells to one line for md/pdf tables."""
    return " ".join(text.split())


def graphql(query: str, variables: dict) -> dict:
    token = os.environ.get("OC_PERSONAL_TOKEN") or os.environ.get("OC_TOKEN")
    headers = {
        "Content-Type": "application/json",
        # Cloudflare fronting the OC API rejects the default Python-urllib
        # user agent with 403 / error 1010.
        "User-Agent": "Mozilla/5.0 (compatible; oc-local-export/1.0)",
    }
    if token:
        headers["Personal-Token"] = token
    body = json.dumps({"query": query, "variables": variables}).encode()
    # The nested transactions/updates/children lookups make pages heavy enough
    # to hit rate limits and timeouts — retry like export_never_active_collectives.py.
    for attempt in range(6):
        req = urllib.request.Request(API_URL, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                payload = json.load(resp)
            break
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 5:
                try:
                    # OC sends a fractional Retry-After (e.g. "16.971").
                    wait = float(e.headers.get("Retry-After") or 0) + 1 or 65
                except ValueError:
                    wait = 65
                print(f"rate limited (429) — waiting {wait}s "
                      f"(set OC_PERSONAL_TOKEN for higher limits)", file=sys.stderr)
                time.sleep(wait)
                continue
            sys.exit(f"HTTP {e.code} from Open Collective: {e.read().decode(errors='replace')}")
        except (urllib.error.URLError, TimeoutError) as e:
            if attempt < 5:
                reason = getattr(e, "reason", e)
                print(f"network error ({reason}) — retrying in 10s", file=sys.stderr)
                time.sleep(10)
                continue
            sys.exit(f"Network error talking to Open Collective: {getattr(e, 'reason', e)}")
    if payload.get("errors"):
        sys.exit("GraphQL errors:\n" + json.dumps(payload["errors"], indent=2))
    return payload.get("data") or {}


def fetch_all(host_slug: str) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    total = None
    while total is None or offset < total:
        data = graphql(QUERY, {"host": [{"slug": host_slug}], "limit": PAGE_SIZE, "offset": offset})
        page = (data.get("accounts") or {})
        total = page.get("totalCount") or 0
        nodes = page.get("nodes") or []
        if not nodes:
            break
        rows.extend(extract_row(c) for c in nodes)
        offset += len(nodes)
        print(f"fetched {min(offset, total)}/{total} collectives...", file=sys.stderr)
    rows.sort(key=lambda r: r["collective"].lower())
    return rows


def to_markers(text: str) -> str:
    """Encode in-cell line breaks as visible markers — spreadsheet importers
    that drop real newlines (e.g. Proton) keep these, and a SUBSTITUTE formula
    turns them back into CHAR(10) line breaks after import."""
    return text.replace("\n\n", " ¶¶ ").replace("\n", " ¶ ")


def write_csv(rows: list[dict], out: str) -> None:
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([label for _, label in COLUMNS])
        for r in rows:
            w.writerow([to_markers(r.get(key, "")) for key, _ in COLUMNS])


def write_xlsx(rows: list[dict], out: str) -> None:
    from openpyxl import Workbook  # declared in the script header
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "collectives"
    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(key, "") for key, _ in COLUMNS])
    widths = {
        "collective": 30, "collective_url": 40, "long_description": 80,
        "tags": 30, "social_links": 45, "categories": 20, "children": 60,
    }
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(key, 16)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(out)


def write_md(rows: list[dict], out: str, title: str) -> None:
    url_cols = set(MD_LINKS.values())
    cols = [(key, label) for key, label in COLUMNS if key not in url_cols]
    labels = [label for _, label in cols]
    lines = [f"# {title}", "", "| " + " | ".join(labels) + " |", "| " + " | ".join(["---"] * len(labels)) + " |"]
    for r in rows:
        cells = []
        for key, _ in cols:
            text = clip(flatten(str(r.get(key, ""))), key, MD_TRUNCATE_AT).replace("|", "\\|")
            url = r.get(MD_LINKS.get(key, ""), "")
            cells.append(f"[{text}]({url})" if url and text else text)
        lines.append("| " + " | ".join(cells) + " |")
    lines += ["", f"_{len(rows)} collective(s)._", ""]
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_pdf(rows: list[dict], out: str, title: str) -> None:
    from fpdf import FPDF  # from fpdf2, declared in the script header

    pdf = FPDF(orientation="L")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    labels = [label for _, label in COLUMNS]
    col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / len(labels)
    pdf.set_font("Helvetica", "B", 8)
    for label in labels:
        pdf.cell(col_w, 8, label, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for r in rows:
        for key, _ in COLUMNS:
            text = clip(flatten(str(r.get(key, ""))), key, PDF_TRUNCATE_AT)
            # latin-1 is fpdf's core-font encoding; drop anything outside it.
            pdf.cell(col_w, 7, text.encode("latin-1", "replace").decode("latin-1"), border=1)
        pdf.ln()
    pdf.output(out)


def run() -> None:
    ap = argparse.ArgumentParser(description="Export a host's collectives with categorization columns to a local file.")
    ap.add_argument("--slug", default="europe", help="Host slug to query (default: europe).")
    ap.add_argument("--format", choices=["csv", "md", "pdf", "xlsx"], default="md")
    ap.add_argument("--out", default=None,
                    help="Output file path (default: output/<host-slug>-collectives.<format> next to this script).")
    args = ap.parse_args()

    if not args.out:
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(out_dir, exist_ok=True)
        args.out = os.path.join(out_dir, f"{args.slug}-collectives.{args.format}")

    rows = fetch_all(args.slug)

    title = f"Collectives — host '{args.slug}'"
    if args.format == "csv":
        write_csv(rows, args.out)
    elif args.format == "xlsx":
        write_xlsx(rows, args.out)
    elif args.format == "md":
        write_md(rows, args.out, title)
    else:
        write_pdf(rows, args.out, title)

    print(f"Wrote {len(rows)} row(s) to {args.out}")


if __name__ == "__main__":
    run()
