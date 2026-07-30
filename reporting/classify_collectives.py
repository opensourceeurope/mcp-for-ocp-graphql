# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Classify the collectives exported by export_host_collectives.py using the
Claude CLI (`claude -p`) — no API key needed, it uses your existing Claude
Code login.

Two modes:

    uv run classify_collectives.py --discover
        Reads the export, asks Claude to propose a category taxonomy from the
        actual data, writes it to output/collectives-taxonomy.md. Review and
        edit that file before classifying.

    uv run classify_collectives.py
        Classifies every collective against the taxonomy file, in batches.
        Appends columns: Open Source (yes/no — does it produce/maintain open
        source software, hardware, or content; advocacy/meetups alone are
        "no"), Tech (yes/no — is the activity technology-centered at all),
        Category 1-3 (from the taxonomy, most specific first),
        Confidence (high/medium/low), AI Reasoning (one sentence).
        Writes <input stem>-classified.csv and .xlsx next to the input.

Per-collective results are cached in output/classify-cache.jsonl, so an
interrupted or repeated run only asks Claude about rows it hasn't judged yet
(--fresh discards the cache). Pass --model to override the CLI's default
model. The Claude CLI must be installed and logged in.
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(HERE, "output", "europe-collectives.csv")
TAXONOMY_PATH = os.path.join(HERE, "output", "collectives-taxonomy.md")
CACHE_PATH = os.path.join(HERE, "output", "classify-cache.jsonl")

NEW_COLUMNS = ["Open Source", "Tech", "Category 1", "Category 2", "Category 3", "Confidence", "AI Reasoning"]
BATCH_SIZE = 15
LONG_DESC_CHARS = 800
PROJECTS_CHARS = 400


def decode_markers(text: str) -> str:
    """The CSV encodes in-cell line breaks as ¶ markers — undo for prompts."""
    return text.replace(" ¶¶ ", "\n").replace(" ¶ ", "\n")


def slug_of(row: dict) -> str:
    return row["Collective URL"].rstrip("/").rsplit("/", 1)[-1]


def ask_claude(prompt: str, model: str | None) -> str:
    cmd = ["claude", "-p", "--output-format", "text"]
    if model:
        cmd += ["--model", model]
    try:
        # Run from a neutral cwd: inside this repo the CLI would load the
        # project's CLAUDE.md and hooks (whose output then pollutes the
        # response — the repo's docs-gate Stop hook, for example).
        r = subprocess.run(cmd, input=prompt, capture_output=True, text=True,
                           timeout=900, cwd=tempfile.gettempdir())
    except FileNotFoundError:
        sys.exit("claude CLI not found — install Claude Code and log in first")
    except subprocess.TimeoutExpired:
        sys.exit("claude CLI timed out after 15 minutes")
    if r.returncode != 0:
        sys.exit(f"claude CLI failed (exit {r.returncode}): {r.stderr.strip()[:500]}")
    return r.stdout


def extract_json_array(text: str) -> list:
    start, end = text.find("["), text.rfind("]")
    if start == -1 or end <= start:
        raise ValueError(f"no JSON array in response: {text[:200]!r}")
    return json.loads(text[start : end + 1])


def collective_blurb(row: dict, full: bool) -> str:
    lines = [f"slug: {slug_of(row)}", f"name: {row['Collective']}"]
    if row["Tags"]:
        lines.append(f"tags: {row['Tags']}")
    if row["Long Description"]:
        limit = LONG_DESC_CHARS if full else 200
        lines.append(f"description: {decode_markers(row['Long Description'])[:limit]}")
    if full and row["Projects"]:
        lines.append(f"sub-projects: {decode_markers(row['Projects'])[:PROJECTS_CHARS]}")
    if full and row["Social Links"]:
        lines.append(f"links: {decode_markers(row['Social Links'])}")
    return "\n".join(lines)


def discover(rows: list[dict], model: str | None) -> None:
    blurbs = "\n\n".join(collective_blurb(r, full=False) for r in rows)
    prompt = f"""You are designing a category taxonomy for the {len(rows)} collectives below,
hosted by a European fiscal host with many (but not only) open source projects.

Propose 12-20 categories that partition this data well. Requirements:
- Categories describe what a collective IS or DOES (e.g. developer tooling,
  communication software, hosting/infrastructure, meetup community, arts).
- Specific enough to be useful for grouping, general enough that each category
  fits several collectives.
- Include exactly one catch-all category named "Other".
- Output ONLY a markdown bullet list, one category per line, in the form:
  - Category Name — one-line definition
No preamble, no explanation after the list.

The collectives:

{blurbs}"""
    print(f"asking Claude to propose a taxonomy from {len(rows)} collectives...", file=sys.stderr)
    answer = ask_claude(prompt, model).strip()
    lines = [line for line in answer.splitlines() if line.lstrip().startswith("- ")]
    if len(lines) < 5:
        sys.exit(f"unexpected taxonomy response:\n{answer[:500]}")
    with open(TAXONOMY_PATH, "w", encoding="utf-8") as f:
        f.write("# Collectives taxonomy\n\n"
                "Edit freely (rename, merge, add, remove) — classification uses\n"
                "exactly the category names listed here.\n\n" + "\n".join(lines) + "\n")
    print(f"Wrote {len(lines)} categories to {TAXONOMY_PATH} — review/edit, then run without --discover")


def load_taxonomy() -> list[str]:
    if not os.path.exists(TAXONOMY_PATH):
        sys.exit(f"{TAXONOMY_PATH} not found — run with --discover first")
    names = []
    for line in open(TAXONOMY_PATH, encoding="utf-8"):
        line = line.strip()
        if line.startswith("- "):
            names.append(line[2:].split(" — ")[0].strip())
    if not names:
        sys.exit(f"no '- Category — definition' lines found in {TAXONOMY_PATH}")
    return names


def classify_batch(batch: list[dict], taxonomy_md: str, categories: list[str], model: str | None) -> dict[str, dict]:
    blurbs = "\n\n---\n\n".join(collective_blurb(r, full=True) for r in batch)
    prompt = f"""Classify each collective below. Use this taxonomy (category names must be
copied EXACTLY from this list):

{taxonomy_md}

For each collective output an object:
- "slug": copied from the input
- "open_source": "yes" only if it produces or maintains open source software,
  hardware, or openly licensed content; advocacy, meetups, user groups, or
  unrelated community projects are "no"
- "tech": "yes" if the collective's activity centers on technology — building
  or operating software, hardware, or digital services/infrastructure, or
  organizing tech communities and events; "no" for arts, mutual aid, local
  community, sports, food, and similar non-technology collectives
- "categories": 1 to 3 category names from the taxonomy, most specific first
- "confidence": "high", "medium" or "low" (how much signal the text gave you)
- "reasoning": one short sentence justifying the judgement

Output ONLY a JSON array with one object per collective, same order as input.

The collectives:

{blurbs}"""
    answer = ask_claude(prompt, model)
    results = {}
    for obj in extract_json_array(answer):
        cats = [c for c in obj.get("categories") or [] if c in categories][:3]
        results[obj["slug"]] = {
            "Open Source": "yes" if obj.get("open_source") == "yes" else "no",
            "Tech": "yes" if obj.get("tech") == "yes" else "no",
            "Category 1": cats[0] if len(cats) > 0 else "",
            "Category 2": cats[1] if len(cats) > 1 else "",
            "Category 3": cats[2] if len(cats) > 2 else "",
            "Confidence": obj.get("confidence") or "",
            "AI Reasoning": (obj.get("reasoning") or "").strip(),
        }
    return results


def load_cache() -> dict[str, dict]:
    cache = {}
    if os.path.exists(CACHE_PATH):
        for line in open(CACHE_PATH, encoding="utf-8"):
            obj = json.loads(line)
            cache[obj["slug"]] = obj["result"]
    return cache


def append_cache(slug: str, result: dict) -> None:
    with open(CACHE_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps({"slug": slug, "result": result}, ensure_ascii=False) + "\n")


def write_outputs(rows: list[dict], header: list[str], out_base: str) -> None:
    with open(out_base + ".csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        w.writerows(rows)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "collectives"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([decode_markers(r.get(k, "")) for k in header])
    widths = {"Collective": 30, "Collective URL": 40, "Long Description": 80, "Tags": 30,
              "Social Links": 45, "Projects": 60, "AI Reasoning": 60}
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, label in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(label, 16)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(out_base + ".xlsx")


def run() -> None:
    ap = argparse.ArgumentParser(description="Judge + categorize exported collectives via the Claude CLI.")
    ap.add_argument("--in", dest="input", default=DEFAULT_INPUT,
                    help=f"Input CSV from export_host_collectives.py (default: {DEFAULT_INPUT})")
    ap.add_argument("--discover", action="store_true",
                    help="Only propose a taxonomy from the data and write it to output/collectives-taxonomy.md.")
    ap.add_argument("--model", default=None, help="Model override passed to `claude --model` (default: CLI default).")
    ap.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    ap.add_argument("--fresh", action="store_true", help="Ignore and rebuild the per-slug result cache.")
    args = ap.parse_args()

    os.makedirs(os.path.join(HERE, "output"), exist_ok=True)
    rows = list(csv.DictReader(open(args.input, encoding="utf-8")))
    if args.discover:
        discover(rows, args.model)
        return

    categories = load_taxonomy()
    taxonomy_md = open(TAXONOMY_PATH, encoding="utf-8").read()
    print(f"taxonomy: {len(categories)} categories from {TAXONOMY_PATH}", file=sys.stderr)

    if args.fresh and os.path.exists(CACHE_PATH):
        os.remove(CACHE_PATH)
    cache = load_cache()
    todo = [r for r in rows if slug_of(r) not in cache]
    print(f"{len(rows)} collectives, {len(rows) - len(todo)} cached, {len(todo)} to classify", file=sys.stderr)

    for i in range(0, len(todo), args.batch_size):
        batch = todo[i : i + args.batch_size]
        results = classify_batch(batch, taxonomy_md, categories, args.model)
        missing = [slug_of(r) for r in batch if slug_of(r) not in results]
        if missing:
            print(f"warning: no result for {missing} — will retry on next run", file=sys.stderr)
        for slug, result in results.items():
            cache[slug] = result
            append_cache(slug, result)
        print(f"classified {min(i + args.batch_size, len(todo))}/{len(todo)}", file=sys.stderr)

    unjudged = 0
    header = list(rows[0].keys()) + NEW_COLUMNS
    for r in rows:
        result = cache.get(slug_of(r))
        if not result:
            unjudged += 1
            result = dict.fromkeys(NEW_COLUMNS, "")
        r.update(result)

    out_base = os.path.splitext(args.input)[0] + "-classified"
    write_outputs(rows, header, out_base)
    note = f" ({unjudged} left unjudged — rerun to retry)" if unjudged else ""
    print(f"Wrote {len(rows)} row(s) to {out_base}.csv and .xlsx{note}")


if __name__ == "__main__":
    run()
